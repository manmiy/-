import json
import re
from pathlib import Path

import streamlit as st
from google.cloud import documentai_v1 as documentai
from google.oauth2 import service_account
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# =====================================================================
# 1. パスワード認証機能
# =====================================================================
def check_password():
    """パスワード認証を行い、認証成功時のみ True を返す"""
    def password_entered():
        correct_password = st.secrets.get("APP_PASSWORD", "password")
        if st.session_state["password_input"] == correct_password:
            st.session_state["password_correct"] = True
            del st.session_state["password_input"]
        else:
            st.session_state["password_correct"] = False

    if st.session_state.get("password_correct", False):
        return True

    st.text_input(
        "パスワードを入力してください",
        type="password",
        on_change=password_entered,
        key="password_input"
    )
    if "password_correct" in st.session_state and not st.session_state["password_correct"]:
        st.error("パスワードが違います")
    return False

# =====================================================================
# 2. Document AI 処理
# =====================================================================
def ocr_pdf_with_documentai(pdf_bytes: bytes, project_id: str, location: str, processor_id: str, credentials):
    """PDFバイトデータをDocument AIに送信し、認識結果を返す"""
    client_options = {"api_endpoint": f"{location}-documentai.googleapis.com"}
    client = documentai.DocumentProcessorServiceClient(
        credentials=credentials, 
        client_options=client_options
    )

    processor_name = client.processor_path(project_id, location, processor_id)
    raw_document = documentai.RawDocument(content=pdf_bytes, mime_type="application/pdf")
    request = documentai.ProcessRequest(name=processor_name, raw_document=raw_document)

    result = client.process_document(request=request)
    return result.document

# =====================================================================
# 3. テキストパース (正規表現)
# =====================================================================
ITEM_PATTERN = re.compile(
    r"^(?P<month>\d{2})\s+(?P<day>\d{2})\s+(?P<denpyo>\d{6})\s+"
    r"(?P<maker>\S+)\s+(?P<hinban>\S+)\s+(?P<hinmei>.+?)\s+"
    r"(?P<qty>-?\d+)\s+(?P<unit>\S+)\s+(?P<price>-?\d+)\s+(?P<amount>-?\d+)"
    r"(?:\s+\((?P<order_no>\d+)\))?"
    r"(?:\s+(?P<note>.+))?$"
)

ITEM_PATTERN_CONT = re.compile(
    r"^(?P<maker>\S+)\s+(?P<hinban>\S+)\s+(?P<hinmei>.+?)\s+"
    r"(?P<qty>-?\d+)\s+(?P<unit>\S+)\s+(?P<price>-?\d+)\s+(?P<amount>-?\d+)"
    r"(?:\s+\((?P<order_no>\d+)\))?"
    r"(?:\s+(?P<note>.+))?$"
)

DENPYO_TOTAL_PATTERN = re.compile(r"^伝票合計\s+(?P<amount>-?\d+)\s*(?:\((?P<order_no>\d+)\))?\s*(?P<note>.+)?$")
KENMEI_TOTAL_PATTERN = re.compile(r"^件名合計\s+(?P<amount>-?\d+)\s*(?:\((?P<order_no>\d+)\))?\s*(?P<note>.+)?$")
TOKUISAKI_TOTAL_PATTERN = re.compile(r"^得意先合計\s+(?P<amount>-?\d+)\s*(?P<note>.+)?$")

def _item_row(date: str, denpyo: str, d: dict) -> dict:
    return {
        "発行日": date, "伝票No": denpyo, "メーカ": d["maker"], "品番": d["hinban"],
        "品名": d["hinmei"], "数量": int(d["qty"]), "単位": d["unit"], "単価": int(d["price"]),
        "金額": int(d["amount"]), "注文No": d.get("order_no") or "", "備考": d.get("note") or "", "行種別": "明細"
    }

def _total_row(date: str, denpyo: str, label: str, d: dict) -> dict:
    return {
        "発行日": date, "伝票No": denpyo, "メーカ": "", "品番": "", "品名": label,
        "数量": "", "単位": "", "単価": "", "金額": int(d["amount"]),
        "注文No": d.get("order_no") or "", "備考": d.get("note") or "", "行種別": label
    }

def parse_lines(ocr_text: str) -> list[dict]:
    rows: list[dict] = []
    current_denpyo, current_date = "", ""

    for raw_line in ocr_text.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        m = ITEM_PATTERN.match(line)
        if m:
            d = m.groupdict()
            current_date, current_denpyo = f"{d['month']}/{d['day']}", d["denpyo"]
            rows.append(_item_row(current_date, current_denpyo, d))
            continue

        m = DENPYO_TOTAL_PATTERN.match(line)
        if m:
            rows.append(_total_row(current_date, current_denpyo, "伝票合計", m.groupdict()))
            continue

        m = KENMEI_TOTAL_PATTERN.match(line)
        if m:
            rows.append(_total_row("", "", "件名合計", m.groupdict()))
            continue

        m = TOKUISAKI_TOTAL_PATTERN.match(line)
        if m:
            rows.append(_total_row("", "", "得意先合計", m.groupdict()))
            continue

        m = ITEM_PATTERN_CONT.match(line)
        if m:
            rows.append(_item_row(current_date, current_denpyo, m.groupdict()))
            continue

    return rows

# =====================================================================
# 4. Excel 生成処理
# =====================================================================
HEADERS = ["発行日", "伝票No", "メーカ", "品番", "品名", "数量", "単位", "単価", "金額", "注文No", "備考"]
COL_WIDTHS = {"発行日": 8, "伝票No": 10, "メーカ": 16, "品番": 14, "品名": 34, "数量": 8, "単位": 6, "単価": 10, "金額": 12, "注文No": 10, "備考": 22}
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def create_excel_bytes(rows: list[dict], sheet_title: str = "4月分請求内訳") -> bytes:
    import io
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    amount_col = HEADERS.index("金額") + 1
    detail_rows_for_sum = []

    row_idx = 2
    for row in rows:
        for col_idx, header in enumerate(HEADERS, start=1):
            value = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial")
            cell.border = BORDER
            if header in ("数量", "単価", "金額") and value != "":
                cell.number_format = "#,##0"
            if row.get("行種別") != "明細":
                cell.fill = TOTAL_FILL
                cell.font = Font(name="Arial", bold=True)
        if row.get("行種別") == "明細":
            detail_rows_for_sum.append(row_idx)
        row_idx += 1

    total_row_idx = row_idx + 1
    ws.cell(row=total_row_idx, column=HEADERS.index("品名") + 1, value="総合計(明細のみ)").font = Font(name="Arial", bold=True)
    col_letter = get_column_letter(amount_col)
    formula = f"=SUM({','.join(f'{col_letter}{r}' for r in detail_rows_for_sum)})" if detail_rows_for_sum else 0
    
    total_cell = ws.cell(row=total_row_idx, column=amount_col, value=formula)
    total_cell.number_format = "#,##0"
    total_cell.font = Font(name="Arial", bold=True)
    total_cell.fill = TOTAL_FILL

    for col_idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(header, 12)

    ws.freeze_panes = "A2"
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# =====================================================================
# 5. Streamlit メインUI画面
# =====================================================================
if check_password():
    st.title("インボイスOCR自動転記システム")

    # サイドバーでGCP設定を入力・管理（Secretsがあれば自動補完）
    st.sidebar.header("Document AI 設定")
    project_id = st.sidebar.text_input("GCP Project ID", value=st.secrets.get("GCP_PROJECT_ID", ""))
    location = st.sidebar.text_input("Location", value=st.secrets.get("GCP_LOCATION", "us"))
    processor_id = st.sidebar.text_input("Processor ID", value=st.secrets.get("GCP_PROCESSOR_ID", ""))

    uploaded_file = st.file_uploader("処理する請求書PDFを選択してください", type=["pdf"])

    if uploaded_file is not None:
        if st.button("OCR処理を開始"):
            if not project_id or not processor_id:
                st.error("サイドバーで Project ID と Processor ID を設定してください。")
            elif "GCP_CREDENTIALS_JSON" not in st.secrets:
                st.error("Streamlit Secrets に 'GCP_CREDENTIALS_JSON' が設定されていません。")
            else:
                with st.spinner("Document AI で解析中..."):
                    try:
                        # 認証情報の読み込み
                        service_account_info = json.loads(st.secrets["GCP_CREDENTIALS_JSON"])
                        credentials = service_account.Credentials.from_service_account_info(service_account_info)

                        # OCR実行
                        pdf_bytes = uploaded_file.read()
                        doc = ocr_pdf_with_documentai(pdf_bytes, project_id, location, processor_id, credentials)

                        # パース処理
                        rows = parse_lines(doc.text)
                        st.success(f"解析完了: {len(rows)} 行の明細を抽出しました。")

                        if rows:
                            # Excelバイナリ生成
                            excel_bytes = create_excel_bytes(rows)
                            st.download_button(
                                label="Excelファイルをダウンロード",
                                data=excel_bytes,
                                file_name="請求内訳.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        else:
                            st.warning("明細行が抽出できませんでした。文字認識結果を確認してください。")
                            with st.expander("OCR解析テキスト（デバッグ用）"):
                                st.text(doc.text)

                    except Exception as e:
                        st.error(f"エラーが発生しました: {e}")
